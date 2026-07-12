#!/usr/bin/env python3
"""
Pushes the e2e Status + Actual Result for every TC into a Google Sheet, using a
Google service account. Source of truth is the local workbook (already updated
by scripts/update-test-results.py), so run that first.

Setup (one time):
  1. In Google Cloud console: create/pick a project, enable the "Google Sheets API".
  2. Create a Service Account -> Keys -> add JSON key -> download it.
  3. Share the Google Sheet (Share button) with the service account's email
     (looks like ...@...iam.gserviceaccount.com) as **Editor**.

Run:
  python3 scripts/push-results-to-gsheet.py \
      --creds /path/to/service-account.json \
      --url "https://docs.google.com/spreadsheets/d/<KEY>/edit?gid=<GID>"

  (or set GOOGLE_APPLICATION_CREDENTIALS instead of --creds)

It locates the worksheet by gid, finds the header row containing "TC ID",
maps the "Status" and "Actual Result" columns by name, and batch-writes only
the rows whose TC ID we have a result for. Nothing else is touched.
"""
import argparse
import os
import re
import sys

import gspread
import openpyxl
from google.oauth2.service_account import Credentials

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Nervaya_Test_Cases_final_version.xlsx")
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


def parse_url(url: str):
    key = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    gid = re.search(r"[#&?]gid=(\d+)", url)
    if not key:
        sys.exit("Could not parse spreadsheet key from --url")
    return key.group(1), (int(gid.group(1)) if gid else None)


def load_results_from_xlsx():
    """Returns {tc_id: (status, actual)} from the local Master Test Cases sheet."""
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Master Test Cases"]
    out = {}
    for r in range(1, ws.max_row + 1):
        tc = ws.cell(r, 1).value
        if isinstance(tc, str) and tc.strip().startswith("TC-"):
            out[tc.strip()] = (ws.cell(r, 8).value or "", ws.cell(r, 7).value or "")
    return out


def col_letter(n: int) -> str:  # 1 -> A
    s = ""
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--creds", default=os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"))
    ap.add_argument("--url", required=True)
    ap.add_argument("--dry-run", action="store_true", help="report what would change, write nothing")
    args = ap.parse_args()
    if not args.creds:
        sys.exit("Provide --creds /path/to/service-account.json (or set GOOGLE_APPLICATION_CREDENTIALS)")

    key, gid = parse_url(args.url)
    results = load_results_from_xlsx()
    print(f"Loaded {len(results)} TC results from {os.path.basename(XLSX)}")

    gc = gspread.authorize(Credentials.from_service_account_file(args.creds, scopes=SCOPES))
    sh = gc.open_by_key(key)
    ws = next((w for w in sh.worksheets() if w.id == gid), None) if gid is not None else sh.sheet1
    if ws is None:
        sys.exit(f"No worksheet with gid={gid} in this spreadsheet. Tabs: {[(w.title, w.id) for w in sh.worksheets()]}")
    print(f"Target worksheet: '{ws.title}' (gid={ws.id})")

    grid = ws.get_all_values()
    # find the header row that contains a "TC ID" cell, and the column indexes.
    header_row = tc_col = status_col = actual_col = None
    for i, row in enumerate(grid):
        norm = [str(c).strip().lower() for c in row]
        if any(c in ("tc id", "tcid", "tc-id") for c in norm):
            header_row = i
            for j, c in enumerate(norm):
                if c in ("tc id", "tcid", "tc-id"):
                    tc_col = j
                elif c == "status":
                    status_col = j
                elif c in ("actual result", "actual"):
                    actual_col = j
            break
    if header_row is None or tc_col is None or status_col is None or actual_col is None:
        sys.exit(f"Could not locate headers (TC ID/Status/Actual Result). Found header_row={header_row}, "
                 f"tc={tc_col}, status={status_col}, actual={actual_col}")
    print(f"Headers @ row {header_row + 1}: TC ID=col{tc_col + 1}, Status=col{status_col + 1}, Actual=col{actual_col + 1}")

    updates, matched, missing = [], 0, []
    for i in range(header_row + 1, len(grid)):
        cell = grid[i][tc_col].strip() if tc_col < len(grid[i]) else ""
        m = re.match(r"(TC-\d+)", cell)
        if not m:
            continue
        tc = m.group(1)
        if tc not in results:
            missing.append(tc)
            continue
        status, actual = results[tc]
        rownum = i + 1
        updates.append({"range": f"{col_letter(status_col + 1)}{rownum}", "values": [[status]]})
        updates.append({"range": f"{col_letter(actual_col + 1)}{rownum}", "values": [[actual]]})
        matched += 1

    print(f"Matched {matched} TC rows in the sheet; {len(missing)} sheet rows had no local result.")
    if args.dry_run:
        print("DRY RUN — no writes. Sample:", updates[:2])
        return
    if updates:
        ws.batch_update(updates, value_input_option="RAW")
    print(f"Done — wrote Status + Actual Result for {matched} test cases to '{ws.title}'.")


if __name__ == "__main__":
    main()
