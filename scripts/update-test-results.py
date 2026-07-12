#!/usr/bin/env python3
"""
Writes Playwright e2e outcomes back into Nervaya_Test_Cases_final_version.xlsx.

- Backs the workbook up to *.bak.xlsx first.
- Reads e2e/.artifacts/results.json, maps each test -> its TC ID(s), and writes
  column G (Actual Result, observed behaviour) + column H (Status).
- TCs that cannot be automated (real WhatsApp/payment/CRM/booking integrations)
  are written as BLOCKED with a concrete reason from BLOCKED_MAP.
- Refreshes the Summary Dashboard counts and appends FAILs to the Bug Log.

Status values written: PASS | FAIL | BLOCKED | REVIEW.
Re-runnable: only TCs the run has info for are touched; human Notes (col J) kept.
"""
import json
import os
import re
import shutil
from datetime import date

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Nervaya_Test_Cases_final_version.xlsx")
BAK = os.path.join(ROOT, "Nervaya_Test_Cases_final_version.bak.xlsx")
RESULTS = os.path.join(ROOT, "e2e/.artifacts/results.json")
LIGHTHOUSE = os.path.join(ROOT, "e2e/.artifacts/lighthouse-summary.json")
STAMP = f"[E2E {date.today().isoformat()}]"

# TCs intentionally NOT automated, with the reason recorded into the sheet.
BLOCKED_MAP = {
    "TC-010": "Requires OTP delivery to a real WhatsApp number; local suite uses the ConsoleOtpDelivery fallback (verified separately). Delivery to real WhatsApp not exercised.",
    "TC-015": "OTP expiry (10-min TTL) not wait-tested in the suite; expiry rejection logic exists (OTP_TTL_MS in otp.constants.ts).",
    "TC-016": "OTP resend has a 10-min (600s) cooldown; the Resend control is present but not exercised to keep the run fast.",
    "TC-017": "Brute-force lockout logic exists (MAX_OTP_VERIFY_ATTEMPTS=5/15min); NOT exercised to avoid locking the shared seeded test accounts.",
    "TC-040": "Needs 3 distinct answer profiles to surface each product type; not automated (single profile exercised).",
    "TC-042": "Retaking the assessment is currently disabled in the app (per register) — nothing to exercise.",
    "TC-048": "Contact form submission not wired to backend/CRM (3rd-party booking + WhatsApp CRM not linked).",
    "TC-049": "Contact form is a placeholder/dummy (per register) — required-field validation not implemented yet.",
    "TC-051": "Enquiry acknowledgement pending (CRM not linked).",
    "TC-060": "Checkout shipping-address validation requires entering Razorpay-gated checkout; blocked by payment integration.",
    "TC-061": "Razorpay card entry happens inside a cross-origin iframe (checkout.razorpay.com) Playwright cannot drive; order-create API is reachable but card submission is not automatable.",
    "TC-062": "Declined-card path is inside the Razorpay cross-origin iframe — not automatable locally.",
    "TC-063": "3DS card flow is inside the Razorpay cross-origin iframe — not automatable locally.",
    "TC-064": "Order confirmation requires a completed (paid) order via the Razorpay iframe — blocked.",
    "TC-065": "Order history requires at least one completed paid order — blocked by payment.",
    "TC-066": "No out-of-stock product exists in the catalog to exercise the out-of-stock UI.",
    "TC-067": "Coupon application happens in the payment-gated checkout — blocked.",
    "TC-068": "Invalid coupon path is in the payment-gated checkout — blocked.",
    "TC-069": "Shipping options shown at the payment-gated checkout — blocked.",
    "TC-070": "Tax calculation shown at the payment-gated checkout — blocked.",
    "TC-071": "Full mobile checkout requires the Razorpay iframe — blocked.",
    "TC-074": "Therapy date/time slot selection is post booking-app integration (per register).",
    "TC-075": "Therapy booking auth gate is post booking-app integration.",
    "TC-076": "Therapy booking payment is post booking-app integration + Razorpay iframe.",
    "TC-077": "Therapy booking confirmation is post booking-app integration.",
    "TC-078": "Double-booking prevention is post booking-app integration.",
    "TC-079": "My Bookings list is post booking-app integration.",
    "TC-080": "Booking cancellation is post booking-app integration.",
    "TC-081": "Booking rescheduling is post booking-app integration.",
    "TC-082": "Booking timezone handling is post booking-app integration.",
    "TC-083": "Calendar invite is post booking-app integration.",
    "TC-085": "Audio purchase requires the Razorpay payment iframe — blocked.",
    "TC-086": "Questionnaire-after-payment requires a completed paid purchase — blocked.",
    "TC-087": "20-question audio form is behind the paid gate — blocked.",
    "TC-088": "Audio questionnaire required-field validation is behind the paid gate — blocked.",
    "TC-089": "Audio questionnaire progress-save is behind the paid gate — blocked.",
    "TC-090": "Audio questionnaire submission is behind the paid gate — blocked.",
    "TC-091": "Audio delivery requires a completed submission behind the paid gate — blocked.",
    "TC-092": "In-browser audio player requires delivered audio (paid) — blocked.",
    "TC-093": "Mobile audio playback requires delivered audio (paid) — blocked.",
    "TC-094": "Re-access audio requires a delivered audio (paid) — blocked.",
    "TC-095": "Payment gate on questionnaire: navigating the questionnaire URL requires payment state — partially design-verified; not exercised.",
    "TC-096": "No-repeat questionnaire requires a submitted questionnaire (paid) — blocked.",
    "TC-097": "Audio delivery SLA requires a real submission/generation — blocked.",
    "TC-098": "Welcome message requires WhatsApp/CRM (CRM pending).",
    "TC-099": "Assessment-results notification requires WhatsApp/CRM (CRM pending).",
    "TC-100": "Order-confirmation message requires WhatsApp/CRM (CRM pending).",
    "TC-101": "Therapy-booking confirmation message requires booking + CRM (pending).",
    "TC-102": "Audio-purchase confirmation message requires payment + CRM (pending).",
    "TC-103": "Audio-ready notification requires CRM (pending).",
    "TC-104": "Contact-enquiry acknowledgement requires CRM (pending).",
    "TC-105": "Spam/deliverability + SPF/DKIM/DMARC require real WhatsApp/email infra (CRM pending).",
    "TC-107": "Razorpay PCI (no raw card data hits Nervaya) cannot be confirmed without driving the payment iframe.",
    "TC-110": "CSRF token inspection requires a state-changing authenticated POST flow; not automated.",
    "TC-111": "OTP/login rate limiting exists (MAX_OTP_SEND_PER_PHONE=5/hr, 429); not flooded in-suite to avoid locking shared accounts.",
    "TC-112": "Razorpay webhook idempotency must be triggered from the Razorpay dashboard — out of scope locally.",
    "TC-113": "Refund flow must be triggered from Razorpay dashboard/admin against a real payment — blocked.",
    "TC-124": "Microsoft Edge engine is not available in this Playwright install (Chromium/Firefox/WebKit only).",
    "TC-125": "Requires a physical iPhone / iOS Safari device (payment + audio).",
    "TC-126": "Requires a physical Android device (payment + audio).",
    "TC-127": "Requires iPad device/emulation for tablet-specific layout; not exercised in this run.",
}

# TCs where the observed reality differs from the register's assumption — recorded as REVIEW.
NOTES_MAP = {
    "TC-007": ("REVIEW", "Page speed/Core Web Vitals must be measured on production (dev/turbopack timings are not representative). See lighthouse-summary if present; register notes homepage LCP > 3.8s."),
    "TC-018": ("REVIEW", "Login does NOT auto-create an account: POST /api/auth/login returns 404 'No account found. Please sign up.' Auto-registration happens via the /signup two-stage flow, not on login."),
    "TC-023": ("REVIEW", "Login-before-assessment: covered indirectly (logged-in user sees results directly, TC-037). Full CTA-from-homepage login path not separately automated."),
    "TC-024": ("REVIEW", "Login-after-assessment: register flags that post-assessment login redirects to dashboard instead of results. Guest gate + logged-in direct results are verified (TC-034/035/037); the gate->login->results hop was not driven end-to-end (OTP at gate)."),
    "TC-036": ("REVIEW", "Post-login result display from the guest gate verified indirectly via logged-in direct results (TC-037); the gate->OTP->results hop not driven end-to-end."),
    "TC-046": ("REVIEW", "Contact option on the results page: support is offered via the left-side panel for logged-in users (per register); contact/support affordance presence verified on homepage/product pages (TC-044/047), results-page-specific affordance not separately asserted."),
    "TC-057": ("REVIEW", "Cart quantity update not reliably automated; register notes only 2 supplements add and values not stored. Cart add/persist/remove are covered by TC-056/058/059."),
}


def parse_tc_ids(title: str):
    ids = re.findall(r"TC-(\d+)", title)
    extra = re.findall(r"/(\d{2,3})(?!\d)", title)  # slash-joined second IDs e.g. TC-034/035
    return [f"TC-{int(x):03d}" for x in ids + extra]


def collect_results():
    """Returns {tc_id: (status, actual)} from the Playwright JSON report."""
    out = {}
    if not os.path.exists(RESULTS):
        print(f"WARNING: {RESULTS} not found")
        return out
    data = json.load(open(RESULTS))

    def status_of(test):
        results = test.get("results", [])
        st = results[-1].get("status") if results else "unknown"
        anns = {a.get("type"): a.get("description") for a in test.get("annotations", [])}
        actual = anns.get("actual")
        if not actual and st != "passed" and results:
            errs = results[-1].get("errors", [])
            if errs:
                actual = re.sub(r"\x1b\[[0-9;]*m", "", errs[0].get("message", "")).splitlines()[0][:200]
        return st, (actual or "")

    def walk(suites):
        for s in suites:
            for sp in s.get("specs", []):
                title = sp.get("title", "")
                fpath = sp.get("file", "") or s.get("file", "")
                is_xbrowser = "11-cross-browser" in fpath or "cross-browser" in title.lower()
                for t in sp.get("tests", []):
                    st, actual = status_of(t)
                    proj = t.get("projectName", "")
                    label = {"passed": "PASS", "failed": "FAIL", "timedOut": "FAIL", "skipped": "REVIEW"}.get(st, "REVIEW")
                    if is_xbrowser:
                        tcid = {"chromium": "TC-121", "firefox": "TC-122", "webkit": "TC-123"}.get(proj)
                        if tcid:
                            merge(out, tcid, label, f"{STAMP} {actual}")
                        continue
                    if proj and proj != "chromium":
                        continue  # non-xbrowser specs only count on chromium
                    for tcid in parse_tc_ids(title):
                        merge(out, tcid, label, f"{STAMP} {actual}")
            walk(s.get("suites", []))

    walk(data.get("suites", []))
    return out


def merge(out, tcid, status, actual):
    # FAIL wins over PASS if a TC maps to multiple tests; keep the most informative actual.
    prio = {"FAIL": 3, "REVIEW": 2, "PASS": 1, "BLOCKED": 0}
    if tcid not in out or prio.get(status, 0) >= prio.get(out[tcid][0], 0):
        out[tcid] = (status, actual)


def main():
    import openpyxl

    shutil.copy2(XLSX, BAK)
    print(f"Backed up workbook -> {BAK}")

    results = collect_results()

    # Lighthouse summary for TC-007, if generated.
    if os.path.exists(LIGHTHOUSE):
        lh = json.load(open(LIGHTHOUSE))
        results["TC-007"] = (lh.get("status", "REVIEW"), f"{STAMP} {lh.get('actual', '')}")

    # Apply NOTES_MAP only where the run produced nothing.
    for tcid, (st, note) in NOTES_MAP.items():
        results.setdefault(tcid, (st, f"{STAMP} {note}"))
    # Apply BLOCKED_MAP only where the run produced nothing.
    for tcid, reason in BLOCKED_MAP.items():
        results.setdefault(tcid, ("BLOCKED", f"{STAMP} Blocked: {reason}"))

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Master Test Cases"]

    # locate header columns (row 2): G=Actual, H=Status
    COL_TC, COL_ACTUAL, COL_STATUS = 1, 7, 8
    written = {"PASS": 0, "FAIL": 0, "BLOCKED": 0, "REVIEW": 0}
    bug_rows = []
    per_module = {}

    for row in range(1, ws.max_row + 1):
        tc = ws.cell(row=row, column=COL_TC).value
        if not isinstance(tc, str) or not tc.strip().startswith("TC-"):
            continue
        tcid = tc.strip()
        module = ws.cell(row=row, column=2).value or ""
        title = ws.cell(row=row, column=4).value or ""
        if tcid in results:
            status, actual = results[tcid]
            ws.cell(row=row, column=COL_ACTUAL).value = actual
            ws.cell(row=row, column=COL_STATUS).value = status
            written[status] = written.get(status, 0) + 1
            per_module.setdefault(module, {"PASS": 0, "FAIL": 0, "BLOCKED": 0, "REVIEW": 0, "TOTAL": 0})
            per_module[module][status] = per_module[module].get(status, 0) + 1
            per_module[module]["TOTAL"] += 1
            if status == "FAIL":
                bug_rows.append((tcid, module, title, actual))

    # Bug Log sheet: append confirmed FAILs.
    if "Bug Log" in wb.sheetnames and bug_rows:
        bl = wb["Bug Log"]
        # find first empty data row (after header row 1)
        r = 2
        while bl.cell(row=r, column=1).value:
            r += 1
        for i, (tcid, module, title, actual) in enumerate(bug_rows, start=1):
            bl.cell(row=r, column=1).value = f"BUG-{i:03d}"
            bl.cell(row=r, column=2).value = tcid
            bl.cell(row=r, column=3).value = module
            bl.cell(row=r, column=4).value = title
            bl.cell(row=r, column=7).value = actual
            bl.cell(row=r, column=8).value = "High"
            bl.cell(row=r, column=9).value = "Open"
            r += 1

    # Refresh the Summary Dashboard (Pass / Fail / Pending=BLOCKED+REVIEW).
    if "Summary Dashboard" in wb.sheetnames:
        ds = wb["Summary Dashboard"]
        dash_map = {
            "General and Navigation": "General",
            "Authentication – WhatsApp OTP": "Auth",
            "Sleep Assessment Flow": "Assessment",
            "Contact and Support": "Contact",
            "E-Commerce – Supplements": "E-Commerce",
            "Therapy Booking": "Therapy",
            "Personalised Audio": "Audio",
            "Email and WhatsApp Notifications": "Notifications",
            "Payment and Security": "Security",
            "Admin and Back-Office": "Admin",
        }
        for row in range(4, 15):
            label = ds.cell(row=row, column=1).value
            if label == "Status Legend":  # repurpose the bogus row for Cross-Browser
                ds.cell(row=row, column=1).value = "Cross-Browser and Device"
                label = "Cross-Browser and Device"
                dash_map[label] = "Cross-Browser"
            mod = dash_map.get(label)
            if not mod or mod not in per_module:
                continue
            c = per_module[mod]
            ds.cell(row=row, column=2).value = c["TOTAL"]
            ds.cell(row=row, column=3).value = c["PASS"]
            ds.cell(row=row, column=4).value = c["FAIL"]
            ds.cell(row=row, column=5).value = c["BLOCKED"] + c["REVIEW"]

    wb.save(XLSX)
    print(f"Updated {sum(written.values())} test cases: {written}")
    print("Per module:")
    for m, c in sorted(per_module.items()):
        print(f"  {m:30s} total={c['TOTAL']:3d}  PASS={c['PASS']} FAIL={c['FAIL']} BLOCKED={c['BLOCKED']} REVIEW={c['REVIEW']}")
    print(f"Bug Log rows appended: {len(bug_rows)}")


if __name__ == "__main__":
    main()
